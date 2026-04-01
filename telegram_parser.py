import asyncio
import re
import json
import os
from datetime import datetime, timezone, timedelta
from telethon import TelegramClient
from telethon.tl.types import MessageMediaDocument, MessageMediaPhoto
import ollama
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ── CONFIG ──────────────────────────────────────────────────────────────────
API_ID    = os.getenv("API_ID")    # set in .env
API_HASH  = os.getenv("API_HASH")  # set in .env
GROUP     = "crossfit"
FRIEND    = "nkabir6202"
OUTPUT       = "/Users/kapilsingla/Documents/crossfit_messages.xlsx"
LAST_RUN     = "/Users/kapilsingla/Documents/telegram-parser/last_run.json"
ATTACHMENTS  = "/Users/kapilsingla/Documents/telegram-attachments"
TWO_YEARS_AGO = datetime(2024, 3, 31, tzinfo=timezone.utc)
MODEL     = "llama3.2"

KNOWN_THEMES = [
    "Educational",
    "AI",
    "Tech Advancement",
    "Politics",
    "Audiobooks / PDF Books",
    "Other",
]
# ────────────────────────────────────────────────────────────────────────────

URL_REGEX = re.compile(r'https?://[^\s]+')

THEME_COLORS = {
    "Educational":          "D9EAD3",
    "AI":                   "CFE2F3",
    "Tech Advancement":     "D0E0E3",
    "Politics":             "FCE5CD",
    "Audiobooks / PDF Books": "EAD1DC",
    "Other":                "F3F3F3",
}

def get_fill(theme: str) -> PatternFill:
    color = THEME_COLORS.get(theme, "FFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def classify_message(text: str, known_themes: list[str]) -> str:
    """Ask Ollama to classify a message into one of the known themes."""
    themes_str = "\n".join(f"- {t}" for t in known_themes)
    prompt = f"""You are a message classifier. Given the message below, pick the SINGLE most fitting theme from the list.
If none fit well, return "Other".
Reply with ONLY the theme name, nothing else.

Themes:
{themes_str}

Message:
{text[:1000]}

Theme:"""
    try:
        response = ollama.chat(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response["message"]["content"].strip().strip('"').strip("'")
        # fuzzy-match to known themes
        for t in known_themes:
            if t.lower() in raw.lower() or raw.lower() in t.lower():
                return t
        return raw if raw else "Other"
    except Exception as e:
        print(f"  [warn] Ollama error: {e}")
        return "Other"


def discover_extra_themes(messages_sample: list[str]) -> list[str]:
    """Ask Ollama to suggest additional themes from a sample of messages."""
    combined = "\n---\n".join(m[:300] for m in messages_sample[:40])
    prompt = f"""Below are messages from a Telegram group. Identify any recurring topics or themes NOT in this list:
{chr(10).join("- " + t for t in KNOWN_THEMES)}

Messages sample:
{combined}

Reply with a JSON array of new theme names only, e.g. ["Health & Fitness", "Finance"].
If there are none, reply with [].
Only theme names, no explanation."""
    try:
        response = ollama.chat(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response["message"]["content"].strip()
        match = re.search(r'\[.*?\]', raw, re.DOTALL)
        if match:
            return json.loads(match.group())
        return []
    except Exception as e:
        print(f"  [warn] Theme discovery error: {e}")
        return []


async def find_group(client, group_name: str):
    """Find a private or public group by title or username."""
    # Try by username first (public groups)
    try:
        return await client.get_entity(group_name)
    except Exception:
        pass
    # Search through all dialogs for a matching title
    print(f"  Searching dialogs for '{group_name}'...")
    async for dialog in client.iter_dialogs():
        if dialog.name and group_name.lower() in dialog.name.lower():
            print(f"  Found: '{dialog.name}' (id={dialog.id})")
            return dialog.entity
    raise ValueError(f"Could not find any group matching '{group_name}'")


def is_conversational(text: str) -> bool:
    """Use Ollama to determine if a message is casual conversation vs a real post/share."""
    prompt = f"""Is the following Telegram message a casual conversational reply (e.g. "ok", "thanks",
"sure", "haha", tagging someone to chat, asking a question in response to someone, small talk)
OR is it a standalone post sharing information, a link, article, book, or resource?

Reply with only one word: "conversation" or "post".

Message: {text[:500]}

Answer:"""
    try:
        response = ollama.chat(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}]
        )
        answer = response["message"]["content"].strip().lower()
        return "conversation" in answer
    except Exception:
        return False


async def fetch_messages(client, group: str, friend: str, since: datetime | None = None) -> list[dict]:
    """Fetch messages from `friend` in `group`, optionally only after `since`."""
    print(f"Looking up group: {group}")
    entity = await find_group(client, group)

    # resolve @mayanknyu to an ID so we can check reply targets
    exclude_reply_to_user = None
    try:
        mayank = await client.get_entity("mayanknyu")
        exclude_reply_to_user = mayank.id
        print(f"  Will exclude replies to @mayanknyu (id={exclude_reply_to_user})")
    except Exception:
        print("  Could not resolve @mayanknyu — skipping that filter")

    os.makedirs(ATTACHMENTS, exist_ok=True)

    # effective start date: whichever is more recent — 2 years ago or last run
    cutoff = TWO_YEARS_AGO
    if since:
        since_utc = since.replace(tzinfo=timezone.utc)
        cutoff = max(cutoff, since_utc)
    print(f"Fetching messages from @{friend} after {cutoff.strftime('%Y-%m-%d')} ...")

    rows = []
    skipped_old = 0
    skipped_no_content = 0
    skipped_reply = 0
    skipped_conversational = 0

    async for msg in client.iter_messages(entity, from_user=friend, limit=None,
                                          offset_date=None,
                                          reverse=False):
        msg_time = msg.date.replace(tzinfo=timezone.utc)

        # stop once we go past the cutoff (messages come newest-first)
        if msg_time <= cutoff:
            skipped_old += 1
            continue

        text = msg.text or ""

        # extract urls from text and entities
        urls = URL_REGEX.findall(text)
        if msg.entities:
            from telethon.tl.types import MessageEntityUrl, MessageEntityTextUrl
            for ent in msg.entities:
                if isinstance(ent, MessageEntityTextUrl):
                    urls.append(ent.url)
                elif isinstance(ent, MessageEntityUrl):
                    urls.append(text[ent.offset: ent.offset + ent.length])
        urls = list(dict.fromkeys(urls))

        # check for attachment
        has_attachment = isinstance(msg.media, MessageMediaDocument)

        # ── Filter 1: must have a link or attachment ──────────────────────────
        if not urls and not has_attachment:
            skipped_no_content += 1
            continue

        # ── Filter 2: skip replies to @mayanknyu ─────────────────────────────
        if exclude_reply_to_user and msg.reply_to:
            try:
                replied_msg = await msg.get_reply_message()
                if replied_msg and replied_msg.sender_id == exclude_reply_to_user:
                    skipped_reply += 1
                    continue
            except Exception:
                pass

        # ── Filter 3: skip pure conversation (Ollama check) ──────────────────
        if text and is_conversational(text):
            skipped_conversational += 1
            print(f"  [skip-conv] {text[:80]}")
            continue

        doc_name = ""
        doc_type = ""
        doc_path = ""
        if has_attachment:
            doc = msg.media.document
            from telethon.tl.types import DocumentAttributeFilename, DocumentAttributeAudio
            for attr in doc.attributes:
                if hasattr(attr, "file_name") and attr.file_name:
                    doc_name = attr.file_name
                if isinstance(attr, DocumentAttributeAudio):
                    doc_type = "Audio"
            if not doc_type:
                mime = getattr(doc, "mime_type", "")
                if "pdf" in mime:
                    doc_type = "PDF"
                elif "audio" in mime:
                    doc_type = "Audio"
                elif "video" in mime:
                    doc_type = "Video"
                else:
                    doc_type = "Document"

            # build a safe filename using msg id + original name
            safe_name = f"{msg.id}_{doc_name}" if doc_name else f"{msg.id}_attachment"
            dest = os.path.join(ATTACHMENTS, safe_name)
            if not os.path.exists(dest):
                print(f"  Downloading: {safe_name}")
                await client.download_media(msg, file=dest)
            else:
                print(f"  Already downloaded: {safe_name}")
            doc_path = dest

        rows.append({
            "date":     msg.date.strftime("%Y-%m-%d %H:%M"),
            "text":     text,
            "urls":     "\n".join(urls),
            "doc_name": doc_name,
            "doc_type": doc_type,
            "doc_path": doc_path,
            "theme":    "",
        })

    print(f"Fetched {len(rows)} messages "
          f"(skipped: {skipped_old} too-old, "
          f"{skipped_no_content} no-link/attachment, "
          f"{skipped_reply} replies-to-mayank, "
          f"{skipped_conversational} conversational)")
    return rows


def build_excel(rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messages by Theme"

    headers = ["Date", "Message", "Theme", "URLs", "Document Name", "Doc Type", "Attachment Link"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # sort by date descending (newest first)
    rows_sorted = sorted(rows, key=lambda r: r["date"], reverse=True)

    link_font = Font(color="0000EE", underline="single")

    for row_idx, r in enumerate(rows_sorted, start=2):
        fill = get_fill(r["theme"])
        values = [r["date"], r["text"], r["theme"], r["urls"], r["doc_name"], r["doc_type"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # attachment hyperlink in column 7
        doc_path = r.get("doc_path", "")
        link_cell = ws.cell(row=row_idx, column=7)
        link_cell.fill = fill
        if doc_path and os.path.exists(doc_path):
            file_uri = "file://" + doc_path
            link_cell.value = os.path.basename(doc_path)
            link_cell.hyperlink = file_uri
            link_cell.font = link_font
        link_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # column widths
    widths = [18, 80, 22, 50, 35, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Theme Summary")
    from collections import Counter
    counts = Counter(r["theme"] for r in rows)
    ws2.append(["Theme", "Message Count"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for theme, count in sorted(counts.items()):
        ws2.append([theme, count])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"\nSaved → {output_path}")


def load_last_run() -> datetime | None:
    """Load the timestamp of the last fetched message."""
    if os.path.exists(LAST_RUN):
        with open(LAST_RUN) as f:
            ts = json.load(f).get("last_message_time")
            if ts:
                return datetime.fromisoformat(ts)
    return None


def save_last_run(ts: datetime):
    """Save the timestamp of the latest message fetched."""
    with open(LAST_RUN, "w") as f:
        json.dump({"last_message_time": ts.isoformat()}, f)


def append_to_excel(rows: list[dict], output_path: str):
    """Append new rows to existing Excel or create new one."""
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Messages by Theme"]
    else:
        build_excel(rows, output_path)
        return

    link_font = Font(color="0000EE", underline="single")
    for r in rows:
        fill = get_fill(r["theme"])
        values = [r["date"], r["text"], r["theme"], r["urls"], r["doc_name"], r["doc_type"]]
        row_idx = ws.max_row + 1
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        doc_path = r.get("doc_path", "")
        link_cell = ws.cell(row=row_idx, column=7)
        link_cell.fill = fill
        if doc_path and os.path.exists(doc_path):
            link_cell.value = os.path.basename(doc_path)
            link_cell.hyperlink = "file://" + doc_path
            link_cell.font = link_font
        link_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # update summary sheet
    from collections import Counter
    ws2 = wb["Theme Summary"]
    ws2.delete_rows(2, ws2.max_row)
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            all_rows.append(row[2])
    counts = Counter(all_rows)
    for theme, count in sorted(counts.items()):
        ws2.append([theme, count])

    wb.save(output_path)
    print(f"Appended {len(rows)} new rows → {output_path}")


async def main():
    last_run = load_last_run()
    if last_run:
        print(f"Incremental run — fetching messages after {last_run.strftime('%Y-%m-%d %H:%M')}")
    else:
        print("First run — fetching all messages")

    async with TelegramClient("session_crossfit", API_ID, API_HASH) as client:
        rows = await fetch_messages(client, GROUP, FRIEND, since=last_run)

        if not rows:
            print("No new messages found.")
            return

        # discover extra themes from a sample
        print("\nDiscovering additional themes with Ollama...")
        texts = [r["text"] for r in rows if r["text"]]
        extra = discover_extra_themes(texts)
        all_themes = KNOWN_THEMES[:-1] + extra + ["Other"]  # keep Other last
        seen = set()
        unique_themes = []
        for t in all_themes:
            if t.lower() not in seen:
                seen.add(t.lower())
                unique_themes.append(t)
        print(f"Themes to use: {unique_themes}")

        # add new theme colors for discovered ones
        palette = ["FFF2CC", "E6D0DE", "D9D2E9", "C9DAF8", "B6D7A8", "FFE599"]
        for i, t in enumerate(extra):
            if t not in THEME_COLORS:
                THEME_COLORS[t] = palette[i % len(palette)]

        # classify each message
        print("\nClassifying messages (this may take a few minutes)...")
        for i, row in enumerate(rows):
            content = row["text"] or row["doc_name"] or row["doc_type"]
            if not content.strip():
                row["theme"] = "Other"
                continue
            row["theme"] = classify_message(content, unique_themes)
            print(f"  [{i+1}/{len(rows)}] {row['theme'][:30]:<30} | {content[:60]}")

        if last_run:
            append_to_excel(rows, OUTPUT)
        else:
            build_excel(rows, OUTPUT)

        # save the timestamp of the most recent message
        latest = max(datetime.fromisoformat(r["date"]) for r in rows)
        save_last_run(latest.replace(tzinfo=timezone.utc))
        print(f"\nDone! Last run saved as {latest.strftime('%Y-%m-%d %H:%M')}")


if __name__ == "__main__":
    asyncio.run(main())
